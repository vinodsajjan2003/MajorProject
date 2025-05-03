import openpyxl
import pandas as pd

# Load the Excel file
try:
    wb = openpyxl.load_workbook('attached_assets/Book1.xlsx')
    print("Successfully loaded the workbook")
    
    # Print sheet names
    print(f"Sheet names: {wb.sheetnames}")
    
    # Read the first sheet
    sheet = wb.active
    
    # Print column headers
    headers = [cell.value for cell in sheet[1]]
    print(f"Column headers: {headers}")
    
    # Print first few rows
    for row_idx in range(2, min(7, sheet.max_row + 1)):
        row_data = [cell.value for cell in sheet[row_idx]]
        print(f"Row {row_idx}: {row_data}")
    
    # Try to read with pandas as well
    print("\nReading with pandas:")
    df = pd.read_excel('attached_assets/Book1.xlsx')
    print("DataFrame shape:", df.shape)
    print("DataFrame columns:", df.columns.tolist())
    print("\nFirst 5 rows:")
    print(df.head())
    
except Exception as e:
    print(f"Error reading Excel file: {str(e)}")